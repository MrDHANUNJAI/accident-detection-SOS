from fastapi import FastAPI, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook

import json
import math
import urllib.parse


# ============================================================
# APPLICATION
# ============================================================

app = FastAPI(
    title="Smart Vehicle SOS System",
    version="3.0.0",
    description="Smart Vehicle SOS monitoring backend for ESP8266/NodeMCU"
)


# ============================================================
# CORS
# ============================================================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)


# ============================================================
# BASE DIRECTORIES
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

DATA_DIR = BASE_DIR / "data"

DATA_DIR.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# FILE PATHS
# ============================================================

INDEX_FILE = BASE_DIR / "index.html"

JSON_FILE = DATA_DIR / "sos_events.json"

EXCEL_FILE = DATA_DIR / "sos_events.xlsx"

SENSOR_FILE = DATA_DIR / "sensor_data.json"

SETTINGS_FILE = DATA_DIR / "settings.json"


# ============================================================
# FALLBACK GPS LOCATION
# ============================================================

FALLBACK_LATITUDE = 13.629000

FALLBACK_LONGITUDE = 78.485000


# ============================================================
# DEFAULT SETTINGS
# ============================================================

DEFAULT_SETTINGS = {
    "vehicleId": "VEHICLE-001",
    "driver": "Driver-01",
    "whatsappPhone": ""
}


# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def now_string():
    """
    Return current server time.
    """

    return datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )


def safe_float(value, default=0.0):
    """
    Convert a value to float safely.
    """

    try:

        number = float(value)

        if math.isfinite(number):
            return number

    except (
        TypeError,
        ValueError
    ):
        pass

    return default


def safe_bool(value, default=False):
    """
    Convert common JSON/string values to bool.
    """

    if isinstance(value, bool):
        return value

    if isinstance(value, str):

        value_lower = value.strip().lower()

        if value_lower in (
            "true",
            "1",
            "yes",
            "on",
            "detected",
            "connected"
        ):
            return True

        if value_lower in (
            "false",
            "0",
            "no",
            "off",
            "not detected",
            "disconnected"
        ):
            return False

    if isinstance(value, (int, float)):

        return bool(value)

    return default


def calculate_magnitude(
    x,
    y,
    z
):
    """
    Calculate acceleration magnitude.
    """

    return math.sqrt(
        x * x +
        y * y +
        z * z
    )


def calculate_impact_g(
    x,
    y,
    z
):
    """
    Calculate impact G after removing
    approximately 1g static baseline.

    If acceleration values are already in g,
    this directly calculates:

        magnitude - 1g
    """

    magnitude = calculate_magnitude(
        x,
        y,
        z
    )

    impact = magnitude - 1.0

    if impact < 0:
        impact = 0.0

    return impact


# ============================================================
# ROOT DASHBOARD
# ============================================================

@app.get(
    "/",
    response_class=HTMLResponse
)
def dashboard():

    if not INDEX_FILE.exists():

        return HTMLResponse(
            content="""
            <!DOCTYPE html>
            <html>
            <head>
                <title>SOS Dashboard Error</title>
            </head>

            <body>

                <h1>Dashboard not found</h1>

                <p>
                    Please place index.html in the
                    same folder as app.py.
                </p>

            </body>
            </html>
            """,
            status_code=500
        )

    return HTMLResponse(
        content=INDEX_FILE.read_text(
            encoding="utf-8"
        )
    )


# ============================================================
# LOAD EVENTS
# ============================================================

def load_events():

    if not JSON_FILE.exists():
        return []

    try:

        with open(
            JSON_FILE,
            "r",
            encoding="utf-8"
        ) as file:

            data = json.load(file)

        if isinstance(data, list):
            return data

        return []

    except Exception as error:

        print(
            "JSON LOAD ERROR:",
            error
        )

        return []


# ============================================================
# SAVE EVENTS
# ============================================================

def save_events(events):

    temporary_file = JSON_FILE.with_suffix(
        ".tmp"
    )

    with open(
        temporary_file,
        "w",
        encoding="utf-8"
    ) as file:

        json.dump(
            events,
            file,
            indent=4,
            ensure_ascii=False
        )

    temporary_file.replace(
        JSON_FILE
    )


# ============================================================
# LOAD SENSOR DATA
# ============================================================

def load_sensor_data():

    if not SENSOR_FILE.exists():
        return []

    try:

        with open(
            SENSOR_FILE,
            "r",
            encoding="utf-8"
        ) as file:

            data = json.load(file)

        if isinstance(data, list):
            return data

        return []

    except Exception as error:

        print(
            "SENSOR JSON LOAD ERROR:",
            error
        )

        return []


# ============================================================
# SAVE SENSOR DATA
# ============================================================

def save_sensor_data(data):

    sensor_data = load_sensor_data()

    sensor_data.append(
        data
    )

    temporary_file = SENSOR_FILE.with_suffix(
        ".tmp"
    )

    with open(
        temporary_file,
        "w",
        encoding="utf-8"
    ) as file:

        json.dump(
            sensor_data,
            file,
            indent=4,
            ensure_ascii=False
        )

    temporary_file.replace(
        SENSOR_FILE
    )


# ============================================================
# GENERATE EVENT ID
# ============================================================

def generate_event_id():

    events = load_events()

    highest = 0

    for event in events:

        event_id = str(
            event.get(
                "eventId",
                ""
            )
        )

        if event_id.startswith(
            "EVT-"
        ):

            try:

                number = int(
                    event_id[4:]
                )

                highest = max(
                    highest,
                    number
                )

            except ValueError:
                pass

    return (
        f"EVT-{highest + 1:06d}"
    )


# ============================================================
# EXCEL HEADERS
# ============================================================

EXCEL_HEADERS = [

    "Event ID",
    "Vehicle ID",
    "Driver",
    "Alert Type",

    "Accel X",
    "Accel Y",
    "Accel Z",

    "Impact G",

    "Acceleration Magnitude",

    "Gyro X",
    "Gyro Y",
    "Gyro Z",

    "Tilt",

    "GPS Latitude",
    "GPS Longitude",
    "GPS Speed km/h",
    "GPS Fix",

    "Location Source",

    "Sensor Source",

    "WiFi Status",
    "WhatsApp Status",

    "Event Status",

    "Cancellation Time",
    "Response Time",

    "Date/Time"
]


# ============================================================
# CREATE EXCEL
# ============================================================

def create_excel():

    if EXCEL_FILE.exists():
        return

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "SOS Events"

    sheet.append(
        EXCEL_HEADERS
    )

    for cell in sheet[1]:

        cell.font = cell.font.copy(
            bold=True
        )

    for column in sheet.columns:

        column_letter = (
            column[0].column_letter
        )

        sheet.column_dimensions[
            column_letter
        ].width = 20

    workbook.save(
        EXCEL_FILE
    )


# ============================================================
# EVENT -> EXCEL ROW
# ============================================================

def event_to_excel_row(event):

    return [

        event.get(
            "eventId",
            ""
        ),

        event.get(
            "vehicleId",
            ""
        ),

        event.get(
            "driver",
            ""
        ),

        event.get(
            "alertType",
            ""
        ),

        event.get(
            "accel_x",
            event.get(
                "accelerationX",
                0
            )
        ),

        event.get(
            "accel_y",
            event.get(
                "accelerationY",
                0
            )
        ),

        event.get(
            "accel_z",
            event.get(
                "accelerationZ",
                0
            )
        ),

        event.get(
            "impact_g",
            0
        ),

        event.get(
            "accelerationMagnitude",
            0
        ),

        event.get(
            "gyro_x",
            event.get(
                "gyroX",
                0
            )
        ),

        event.get(
            "gyro_y",
            event.get(
                "gyroY",
                0
            )
        ),

        event.get(
            "gyro_z",
            event.get(
                "gyroZ",
                0
            )
        ),

        event.get(
            "tilt",
            0
        ),

        event.get(
            "gps_lat",
            event.get(
                "latitude",
                FALLBACK_LATITUDE
            )
        ),

        event.get(
            "gps_lon",
            event.get(
                "longitude",
                FALLBACK_LONGITUDE
            )
        ),

        event.get(
            "gps_speed_kmph",
            0
        ),

        event.get(
            "gps_fix",
            event.get(
                "gpsDetected",
                False
            )
        ),

        event.get(
            "locationSource",
            "FALLBACK"
        ),

        event.get(
            "sensorSource",
            "UNKNOWN"
        ),

        event.get(
            "wifiStatus",
            "UNKNOWN"
        ),

        event.get(
            "whatsappStatus",
            "NOT_CONFIGURED"
        ),

        event.get(
            "eventStatus",
            "ACTIVE"
        ),

        event.get(
            "cancellationTime",
            ""
        ),

        event.get(
            "responseTime",
            ""
        ),

        event.get(
            "dateTime",
            ""
        )
    ]


# ============================================================
# SAVE EVENT TO EXCEL
# ============================================================

def save_event_to_excel(event):

    create_excel()

    workbook = load_workbook(
        EXCEL_FILE
    )

    sheet = workbook[
        "SOS Events"
    ]

    sheet.append(
        event_to_excel_row(
            event
        )
    )

    workbook.save(
        EXCEL_FILE
    )


# ============================================================
# UPDATE EVENT IN EXCEL
# ============================================================

def update_event_in_excel(event):

    create_excel()

    workbook = load_workbook(
        EXCEL_FILE
    )

    sheet = workbook[
        "SOS Events"
    ]

    event_id = event.get(
        "eventId"
    )

    found = False

    for row in sheet.iter_rows(
        min_row=2
    ):

        if row[0].value == event_id:

            values = event_to_excel_row(
                event
            )

            for index, value in enumerate(
                values,
                start=1
            ):

                sheet.cell(
                    row=row[0].row,
                    column=index
                ).value = value

            found = True

            break

    if not found:

        sheet.append(
            event_to_excel_row(
                event
            )
        )

    workbook.save(
        EXCEL_FILE
    )


# ============================================================
# HEALTH API
# ============================================================

@app.get(
    "/api/health"
)
def health():

    events = load_events()

    return {

        "status": "ok",

        "message":
            "Smart Vehicle SOS Server is running",

        "database":
            "Excel + JSON",

        "excelExists":
            EXCEL_FILE.exists(),

        "jsonExists":
            JSON_FILE.exists(),

        "sensorFileExists":
            SENSOR_FILE.exists(),

        "totalEvents":
            len(events),

        "timestamp":
            now_string()
    }


# ============================================================
# GET ALL EVENTS
# ============================================================

@app.get(
    "/api/events"
)
def get_events():

    events = load_events()

    # Newest event first.
    events = list(
        reversed(events)
    )

    return {

        "status":
            "success",

        "count":
            len(events),

        "events":
            events
    }


# ============================================================
# NORMALIZE SENSOR/GPS INPUT
# ============================================================

def normalize_sensor_data(data):

    # --------------------------------------------------------
    # ACCELERATION
    # --------------------------------------------------------

    accel_x = safe_float(
        data.get(
            "accel_x",
            data.get(
                "accelerationX",
                0
            )
        )
    )

    accel_y = safe_float(
        data.get(
            "accel_y",
            data.get(
                "accelerationY",
                0
            )
        )
    )

    accel_z = safe_float(
        data.get(
            "accel_z",
            data.get(
                "accelerationZ",
                0
            )
        )
    )

    # --------------------------------------------------------
    # ACCELERATION MAGNITUDE
    # --------------------------------------------------------

    magnitude = calculate_magnitude(
        accel_x,
        accel_y,
        accel_z
    )

    # --------------------------------------------------------
    # IMPACT G
    # --------------------------------------------------------

    if "impact_g" in data:

        impact_g = safe_float(
            data.get(
                "impact_g"
            )
        )

    else:

        impact_g = calculate_impact_g(
            accel_x,
            accel_y,
            accel_z
        )

    # --------------------------------------------------------
    # GYROSCOPE
    # --------------------------------------------------------

    gyro_x = safe_float(
        data.get(
            "gyro_x",
            data.get(
                "gyroX",
                0
            )
        )
    )

    gyro_y = safe_float(
        data.get(
            "gyro_y",
            data.get(
                "gyroY",
                0
            )
        )
    )

    gyro_z = safe_float(
        data.get(
            "gyro_z",
            data.get(
                "gyroZ",
                0
            )
        )
    )

    # --------------------------------------------------------
    # GPS
    # --------------------------------------------------------

    gps_fix = safe_bool(
        data.get(
            "gps_fix",
            data.get(
                "gpsDetected",
                False
            )
        )
    )

    gps_lat = safe_float(
        data.get(
            "gps_lat",
            data.get(
                "latitude",
                FALLBACK_LATITUDE
            )
        ),
        FALLBACK_LATITUDE
    )

    gps_lon = safe_float(
        data.get(
            "gps_lon",
            data.get(
                "longitude",
                FALLBACK_LONGITUDE
            )
        ),
        FALLBACK_LONGITUDE
    )

    gps_speed_kmph = safe_float(
        data.get(
            "gps_speed_kmph",
            data.get(
                "gpsSpeedKmph",
                data.get(
                    "gpsSpeed",
                    0
                )
            )
        )
    )

    # --------------------------------------------------------
    # LOCATION SOURCE
    # --------------------------------------------------------

    if gps_fix:

        location_source = "GPS"

    else:

        gps_lat = FALLBACK_LATITUDE
        gps_lon = FALLBACK_LONGITUDE

        location_source = "FALLBACK"

    # --------------------------------------------------------
    # TILT
    # --------------------------------------------------------

    tilt = safe_float(
        data.get(
            "tilt",
            0
        )
    )

    # --------------------------------------------------------
    # SOURCE
    # --------------------------------------------------------

    sensor_source = data.get(
        "sensorSource",
        "MPU6050"
    )

    wifi_status = data.get(
        "wifiStatus",
        "CONNECTED"
    )

    return {

        "accel_x":
            accel_x,

        "accel_y":
            accel_y,

        "accel_z":
            accel_z,

        "impact_g":
            impact_g,

        "accelerationMagnitude":
            magnitude,

        "gyro_x":
            gyro_x,

        "gyro_y":
            gyro_y,

        "gyro_z":
            gyro_z,

        "tilt":
            tilt,

        "gps_lat":
            gps_lat,

        "gps_lon":
            gps_lon,

        "gps_speed_kmph":
            gps_speed_kmph,

        "gps_fix":
            gps_fix,

        "locationSource":
            location_source,

        "sensorSource":
            sensor_source,

        "wifiStatus":
            wifi_status
    }


# ============================================================
# RECEIVE SOS
# ============================================================

@app.post(
    "/api/sos"
)
async def receive_sos(
    request: Request
):

    try:

        data = await request.json()

        if not isinstance(
            data,
            dict
        ):

            raise HTTPException(
                status_code=400,
                detail="SOS data must be a JSON object"
            )

        print()
        print(
            "=========================================="
        )

        print(
            "       NEW SOS ALERT RECEIVED"
        )

        print(
            "=========================================="
        )

        print(
            "DATA:",
            data
        )

        # ----------------------------------------------------
        # SETTINGS
        # ----------------------------------------------------

        settings = load_settings()

        # ----------------------------------------------------
        # EVENT ID
        # ----------------------------------------------------

        event_id = generate_event_id()

        # ----------------------------------------------------
        # DATE/TIME
        # ----------------------------------------------------

        date_time = now_string()

        # ----------------------------------------------------
        # NORMALIZE SENSOR/GPS
        # ----------------------------------------------------

        sensor = normalize_sensor_data(
            data
        )

        # ----------------------------------------------------
        # VEHICLE
        # ----------------------------------------------------

        vehicle_id = data.get(
            "vehicleId",
            settings.get(
                "vehicleId",
                "VEHICLE-001"
            )
        )

        driver = data.get(
            "driver",
            settings.get(
                "driver",
                "Driver-01"
            )
        )

        # ----------------------------------------------------
        # ALERT TYPE
        # ----------------------------------------------------

        alert_type = data.get(
            "alertType",
            data.get(
                "event",
                "MANUAL_SOS"
            )
        )

        # ----------------------------------------------------
        # WHATSAPP STATUS
        # ----------------------------------------------------

        whatsapp_status = data.get(
            "whatsappStatus",
            "PENDING"
        )

        # ----------------------------------------------------
        # CREATE EVENT
        # ----------------------------------------------------

        event = {

            "eventId":
                event_id,

            "vehicleId":
                str(vehicle_id),

            "driver":
                str(driver),

            "alertType":
                str(alert_type),

            # New exact sensor fields
            "accel_x":
                sensor["accel_x"],

            "accel_y":
                sensor["accel_y"],

            "accel_z":
                sensor["accel_z"],

            "gyro_x":
                sensor["gyro_x"],

            "gyro_y":
                sensor["gyro_y"],

            "gyro_z":
                sensor["gyro_z"],

            "impact_g":
                sensor["impact_g"],

            "gps_lat":
                sensor["gps_lat"],

            "gps_lon":
                sensor["gps_lon"],

            "gps_speed_kmph":
                sensor["gps_speed_kmph"],

            "gps_fix":
                sensor["gps_fix"],

            # Dashboard compatibility fields
            "accelerationX":
                sensor["accel_x"],

            "accelerationY":
                sensor["accel_y"],

            "accelerationZ":
                sensor["accel_z"],

            "accelerationMagnitude":
                sensor["accelerationMagnitude"],

            "gyroX":
                sensor["gyro_x"],

            "gyroY":
                sensor["gyro_y"],

            "gyroZ":
                sensor["gyro_z"],

            "tilt":
                sensor["tilt"],

            "latitude":
                sensor["gps_lat"],

            "longitude":
                sensor["gps_lon"],

            "gpsSpeedKmph":
                sensor["gps_speed_kmph"],

            "gpsDetected":
                sensor["gps_fix"],

            "locationSource":
                sensor["locationSource"],

            "sensorSource":
                sensor["sensorSource"],

            "wifiStatus":
                sensor["wifiStatus"],

            "whatsappStatus":
                whatsapp_status,

            "eventStatus":
                "ACTIVE",

            "cancellationTime":
                "",

            "responseTime":
                "",

            "dateTime":
                date_time
        }

        # ----------------------------------------------------
        # SAVE JSON
        # ----------------------------------------------------

        events = load_events()

        events.append(
            event
        )

        save_events(
            events
        )

        # ----------------------------------------------------
        # SAVE EXCEL
        # ----------------------------------------------------

        save_event_to_excel(
            event
        )

        # ----------------------------------------------------
        # CONSOLE
        # ----------------------------------------------------

        print(
            "EVENT SAVED:",
            event_id
        )

        print(
            "VEHICLE:",
            vehicle_id
        )

        print(
            "ALERT:",
            alert_type
        )

        print(
            "IMPACT G:",
            sensor["impact_g"]
        )

        print(
            "GPS FIX:",
            sensor["gps_fix"]
        )

        print(
            "LOCATION:",
            sensor["gps_lat"],
            sensor["gps_lon"]
        )

        print(
            "STATUS:",
            "ACTIVE"
        )

        print(
            "=========================================="
        )

        # ----------------------------------------------------
        # RESPONSE
        # ----------------------------------------------------

        return {

            "status":
                "success",

            "message":
                "SOS received and stored successfully",

            "event":
                event
        }

    except HTTPException:
        raise

    except Exception as error:

        print(
            "SOS ERROR:",
            error
        )

        raise HTTPException(
            status_code=500,
            detail=str(error)
        )


# ============================================================
# SENSOR DATA
# ============================================================

@app.post(
    "/api/sensor"
)
async def receive_sensor(
    request: Request
):

    try:

        data = await request.json()

        if not isinstance(
            data,
            dict
        ):

            raise HTTPException(
                status_code=400,
                detail="Sensor data must be a JSON object"
            )

        # ----------------------------------------------------
        # Normalize data
        # ----------------------------------------------------

        normalized = normalize_sensor_data(
            data
        )

        # ----------------------------------------------------
        # Preserve extra device fields
        # ----------------------------------------------------

        sensor_record = {

            "timestamp":
                now_string(),

            **normalized
        }

        # ----------------------------------------------------
        # Optional device information
        # ----------------------------------------------------

        if "vehicleId" in data:

            sensor_record[
                "vehicleId"
            ] = data[
                "vehicleId"
            ]

        if "deviceId" in data:

            sensor_record[
                "deviceId"
            ] = data[
                "deviceId"
            ]

        if "uptime" in data:

            sensor_record[
                "uptime"
            ] = data[
                "uptime"
            ]

        # ----------------------------------------------------
        # SAVE
        # ----------------------------------------------------

        save_sensor_data(
            sensor_record
        )

        return {

            "status":
                "success",

            "message":
                "Sensor data received",

            "data":
                sensor_record
        }

    except HTTPException:
        raise

    except Exception as error:

        print(
            "SENSOR ERROR:",
            error
        )

        raise HTTPException(
            status_code=500,
            detail=str(error)
        )


# ============================================================
# GET SENSOR DATA
# ============================================================

@app.get(
    "/api/sensor"
)
def get_sensor():

    return {

        "status":
            "success",

        "count":
            len(load_sensor_data()),

        "data":
            load_sensor_data()
    }


# ============================================================
# GET LATEST SENSOR DATA
# ============================================================

@app.get(
    "/api/sensor/latest"
)
def get_latest_sensor():

    sensor_data = load_sensor_data()

    if not sensor_data:

        return {

            "status":
                "success",

            "data":
                None
        }

    return {

        "status":
            "success",

        "data":
            sensor_data[-1]
    }


# ============================================================
# GET SINGLE EVENT
# ============================================================

@app.get(
    "/api/events/{event_id}"
)
def get_single_event(
    event_id: str
):

    events = load_events()

    for event in events:

        if event.get(
            "eventId"
        ) == event_id:

            return {

                "status":
                    "success",

                "event":
                    event
            }

    raise HTTPException(
        status_code=404,
        detail="Event not found"
    )


# ============================================================
# CANCEL EVENT
# ============================================================

@app.patch(
    "/api/events/{event_id}/cancel"
)
def cancel_event(
    event_id: str
):

    events = load_events()

    for event in events:

        if event.get(
            "eventId"
        ) == event_id:

            if event.get(
                "eventStatus"
            ) == "CANCELLED":

                return {

                    "status":
                        "success",

                    "message":
                        "SOS is already cancelled",

                    "event":
                        event
                }

            event[
                "eventStatus"
            ] = "CANCELLED"

            event[
                "cancellationTime"
            ] = now_string()

            event[
                "responseTime"
            ] = now_string()

            save_events(
                events
            )

            update_event_in_excel(
                event
            )

            print(
                "EVENT CANCELLED:",
                event_id
            )

            return {

                "status":
                    "success",

                "message":
                    "SOS cancelled",

                "event":
                    event
            }

    raise HTTPException(
        status_code=404,
        detail="Event not found"
    )


# ============================================================
# RESOLVE EVENT
# ============================================================

@app.patch(
    "/api/events/{event_id}/resolve"
)
def resolve_event(
    event_id: str
):

    events = load_events()

    for event in events:

        if event.get(
            "eventId"
        ) == event_id:

            if event.get(
                "eventStatus"
            ) == "RESOLVED":

                return {

                    "status":
                        "success",

                    "message":
                        "SOS is already resolved",

                    "event":
                        event
                }

            event[
                "eventStatus"
            ] = "RESOLVED"

            event[
                "responseTime"
            ] = now_string()

            save_events(
                events
            )

            update_event_in_excel(
                event
            )

            print(
                "EVENT RESOLVED:",
                event_id
            )

            return {

                "status":
                    "success",

                "message":
                    "SOS resolved",

                "event":
                    event
            }

    raise HTTPException(
        status_code=404,
        detail="Event not found"
    )


# ============================================================
# SETTINGS
# ============================================================

def load_settings():

    if not SETTINGS_FILE.exists():

        return DEFAULT_SETTINGS.copy()

    try:

        with open(
            SETTINGS_FILE,
            "r",
            encoding="utf-8"
        ) as file:

            data = json.load(
                file
            )

        if not isinstance(
            data,
            dict
        ):

            return DEFAULT_SETTINGS.copy()

        settings = DEFAULT_SETTINGS.copy()

        settings.update(
            data
        )

        return settings

    except Exception as error:

        print(
            "SETTINGS LOAD ERROR:",
            error
        )

        return DEFAULT_SETTINGS.copy()


# ============================================================
# SAVE SETTINGS
# ============================================================

def save_settings(
    settings
):

    with open(
        SETTINGS_FILE,
        "w",
        encoding="utf-8"
    ) as file:

        json.dump(
            settings,
            file,
            indent=4,
            ensure_ascii=False
        )


# ============================================================
# GET SETTINGS
# ============================================================

@app.get(
    "/api/settings"
)
def get_settings():

    return load_settings()


# ============================================================
# UPDATE SETTINGS
# ============================================================

@app.put(
    "/api/settings"
)
async def update_settings(
    request: Request
):

    try:

        data = await request.json()

        if not isinstance(
            data,
            dict
        ):

            raise HTTPException(
                status_code=400,
                detail="Settings must be a JSON object"
            )

        settings = load_settings()

        allowed_fields = {

            "vehicleId",
            "driver",
            "whatsappPhone"
        }

        for key, value in data.items():

            if key in allowed_fields:

                settings[key] = str(
                    value
                ).strip()

        save_settings(
            settings
        )

        return {

            "status":
                "success",

            "message":
                "Settings updated",

            "settings":
                settings
        }

    except HTTPException:
        raise

    except Exception as error:

        raise HTTPException(
            status_code=500,
            detail=str(error)
        )


# ============================================================
# WHATSAPP MESSAGE
# ============================================================

@app.get(
    "/api/whatsapp/{event_id}"
)
def whatsapp_event(
    event_id: str
):

    events = load_events()

    settings = load_settings()

    phone = str(
        settings.get(
            "whatsappPhone",
            ""
        )
    ).strip()

    for event in events:

        if event.get(
            "eventId"
        ) == event_id:

            latitude = event.get(
                "gps_lat",
                event.get(
                    "latitude",
                    FALLBACK_LATITUDE
                )
            )

            longitude = event.get(
                "gps_lon",
                event.get(
                    "longitude",
                    FALLBACK_LONGITUDE
                )
            )

            message = (

                "SMART VEHICLE SOS ALERT\n\n"

                f"Event ID: "
                f"{event.get('eventId')}\n"

                f"Vehicle: "
                f"{event.get('vehicleId')}\n"

                f"Driver: "
                f"{event.get('driver')}\n"

                f"Alert Type: "
                f"{event.get('alertType')}\n"

                f"Status: "
                f"{event.get('eventStatus')}\n\n"

                f"Impact G: "
                f"{event.get('impact_g', 0)}\n"

                f"GPS Fix: "
                f"{event.get('gps_fix', False)}\n"

                f"Location: "
                f"{latitude}, "
                f"{longitude}\n"

                f"GPS Speed: "
                f"{event.get('gps_speed_kmph', 0)} km/h\n"

                f"Time: "
                f"{event.get('dateTime')}"
            )

            encoded_message = (
                urllib.parse.quote(
                    message
                )
            )

            if phone:

                url = (
                    "https://wa.me/"
                    f"{phone}"
                    "?text="
                    f"{encoded_message}"
                )

            else:

                url = (
                    "https://wa.me/"
                    "?text="
                    f"{encoded_message}"
                )

            return {

                "status":
                    "success",

                "eventId":
                    event_id,

                "phoneConfigured":
                    bool(phone),

                "url":
                    url
            }

    raise HTTPException(
        status_code=404,
        detail="Event not found"
    )


# ============================================================
# EXPORT EXCEL
# ============================================================

@app.get(
    "/api/export/excel"
)
def export_excel():

    create_excel()

    return FileResponse(

        path=str(
            EXCEL_FILE
        ),

        filename=
            "sos_events.xlsx",

        media_type=
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
    )


# ============================================================
# DELETE ALL EVENTS
# ============================================================

@app.delete(
    "/api/events"
)
def delete_all_events():

    save_events([])

    # Recreate Excel file.
    if EXCEL_FILE.exists():

        EXCEL_FILE.unlink()

    create_excel()

    return {

        "status":
            "success",

        "message":
            "All SOS events deleted"
    }


# ============================================================
# STARTUP
# ============================================================

@app.on_event(
    "startup"
)
def startup():

    create_excel()

    if not JSON_FILE.exists():

        save_events([])

    if not SENSOR_FILE.exists():

        with open(
            SENSOR_FILE,
            "w",
            encoding="utf-8"
        ) as file:

            json.dump(
                [],
                file,
                indent=4
            )

    if not SETTINGS_FILE.exists():

        save_settings(
            DEFAULT_SETTINGS.copy()
        )

    print()
    print(
        "=========================================="
    )

    print(
        "       SMART VEHICLE SOS SYSTEM"
    )

    print(
        "=========================================="
    )

    print(
        "Dashboard : "
        "http://127.0.0.1:8001/"
    )

    print(
        "Health    : "
        "http://127.0.0.1:8001/api/health"
    )

    print(
        "SOS       : "
        "http://127.0.0.1:8001/api/sos"
    )

    print(
        "Events    : "
        "http://127.0.0.1:8001/api/events"
    )

    print(
        "Sensor    : "
        "http://127.0.0.1:8001/api/sensor"
    )

    print(
        "Settings  : "
        "http://127.0.0.1:8001/api/settings"
    )

    print(
        "Excel     : "
        "http://127.0.0.1:8001/api/export/excel"
    )

    print(
        "=========================================="
    )


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    import uvicorn

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8001,
        reload=False
    )
